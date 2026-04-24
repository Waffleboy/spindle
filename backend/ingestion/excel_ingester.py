"""Excel (.xlsx / .xls) document ingesters using openpyxl and xlrd."""

from pathlib import Path

import xlrd
from openpyxl import load_workbook

from backend.ingestion.common import (
    BaseIngester,
    IngestedDocument,
    register_ingester,
)


class ExcelIngester(BaseIngester):
    """Extracts structured text from Excel (.xlsx) files."""

    def ingest(self, file_path: Path, storage_path: str) -> IngestedDocument:
        """Ingest an Excel workbook.

        Reads all sheets and converts to a structured text representation
        with tab-separated values. page_count equals the number of sheets.

        Args:
            file_path: Path to the .xlsx file.
            storage_path: Relative storage path for reference.

        Returns:
            IngestedDocument with structured text and empty pages list.
        """
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        try:
            text_parts: list[str] = []
            sheet_names: list[str] = []

            for sheet_name in wb.sheetnames:
                sheet_names.append(sheet_name)
                ws = wb[sheet_name]
                sheet_lines: list[str] = [f"Sheet: {sheet_name}"]

                for row in ws.iter_rows(values_only=True):
                    cell_values = [
                        str(cell) if cell is not None else "" for cell in row
                    ]
                    # Skip entirely empty rows
                    if any(v for v in cell_values):
                        sheet_lines.append("\t".join(cell_values))

                text_parts.append("\n".join(sheet_lines))

            text = "\n\n".join(text_parts)

            return IngestedDocument(
                original_filename=file_path.name,
                storage_path=storage_path,
                file_type="xlsx",
                pages=[],
                text=text,
                metadata={"sheet_names": sheet_names},
                page_count=len(wb.sheetnames),
            )
        finally:
            wb.close()


register_ingester("xlsx", ExcelIngester)


class XlsIngester(BaseIngester):
    """Extracts structured text from legacy Excel (.xls) files using xlrd."""

    def ingest(self, file_path: Path, storage_path: str) -> IngestedDocument:
        """Ingest a legacy .xls workbook.

        Reads all sheets and converts to a structured text representation
        with tab-separated values. page_count equals the number of sheets.

        Args:
            file_path: Path to the .xls file.
            storage_path: Relative storage path for reference.

        Returns:
            IngestedDocument with structured text and empty pages list.
        """
        wb = xlrd.open_workbook(str(file_path))
        text_parts: list[str] = []
        sheet_names: list[str] = []

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            sheet_names.append(ws.name)
            sheet_lines: list[str] = [f"Sheet: {ws.name}"]

            for row_idx in range(ws.nrows):
                cell_values = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        cell_values.append("")
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        # Render integers without trailing .0
                        val = cell.value
                        cell_values.append(
                            str(int(val)) if val == int(val) else str(val)
                        )
                    else:
                        cell_values.append(str(cell.value))

                # Skip entirely empty rows
                if any(v for v in cell_values):
                    sheet_lines.append("\t".join(cell_values))

            text_parts.append("\n".join(sheet_lines))

        text = "\n\n".join(text_parts)

        return IngestedDocument(
            original_filename=file_path.name,
            storage_path=storage_path,
            file_type="xls",
            pages=[],
            text=text,
            metadata={"sheet_names": sheet_names},
            page_count=wb.nsheets,
        )


register_ingester("xls", XlsIngester)
